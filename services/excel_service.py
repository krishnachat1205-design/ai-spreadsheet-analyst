"""
Excel Export Service
====================

Generates professionally formatted Excel workbooks containing:
- Cleaned data
- KPI summaries
- Business insights
- Recommendations
- Analyst report

Uses openpyxl for formatting.
"""

from __future__ import annotations

import logging
from datetime import datetime, timezone
from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from services.insight_service import Insight, Recommendation
from services.kpi_service import KPIResult
from services.report_service import AnalystReport

logger = logging.getLogger(__name__)

# =============================================================================
# Styling Constants
# =============================================================================

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGNMENT = Alignment(vertical="center", wrap_text=False)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
TITLE_FONT = Font(bold=True, size=14, color="366092")
SUBTITLE_FONT = Font(bold=True, size=11, color="000000")
METRIC_FONT = Font(bold=True, size=12, color="366092")
NOTE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
CRITICAL_FILL = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

# Maximum rows to write per sheet (safety limit for large datasets)
MAX_EXPORT_ROWS = 100_000


def _auto_column_widths(worksheet) -> None:
    """Auto-adjust column widths based on content."""
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 60)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def _apply_header_style(worksheet, row_idx: int = 1) -> None:
    """Apply header styling to a specific row."""
    for cell in worksheet[row_idx]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _apply_data_borders(worksheet, start_row: int = 2) -> None:
    """Apply thin borders to all data cells."""
    for row in worksheet.iter_rows(min_row=start_row):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = DATA_ALIGNMENT


def _write_dataframe_to_sheet(
    worksheet,
    df: pd.DataFrame,
    start_row: int = 1,
    start_col: int = 1,
    apply_header: bool = True,
    apply_borders: bool = True,
) -> int:
    """
    Write a DataFrame to a worksheet and return the last row index.
    Respects MAX_EXPORT_ROWS limit.
    """
    rows_to_write = min(len(df), MAX_EXPORT_ROWS)
    df_write = df.head(rows_to_write)

    for r_idx, row in enumerate(dataframe_to_rows(df_write, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, start_col):
            worksheet.cell(row=r_idx, column=c_idx, value=value)

    if apply_header:
        _apply_header_style(worksheet, row_idx=start_row)
    if apply_borders:
        _apply_data_borders(worksheet, start_row=start_row + 1)

    return start_row + len(df_write)


# =============================================================================
# Sheet Builders
# =============================================================================


def _build_cleaned_data_sheet(workbook: Workbook, df: pd.DataFrame) -> None:
    """Build Sheet 1: Cleaned Data."""
    ws = workbook.active
    ws.title = "Cleaned Data"

    # Title
    ws["A1"] = "Cleaned Dataset"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    ws["A2"] = f"Exported: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    ws["A2"].font = Font(italic=True, size=9, color="666666")
    ws.merge_cells("A2:D2")

    ws["A3"] = f"Rows: {len(df):,} | Columns: {len(df.columns)}"
    ws["A3"].font = Font(size=9, color="666666")
    ws.merge_cells("A3:D3")

    # Data starting at row 5
    last_row = _write_dataframe_to_sheet(ws, df, start_row=5)

    # Freeze panes below header
    ws.freeze_panes = "A6"

    _auto_column_widths(ws)

    # Add truncation note if dataset was limited
    if len(df) > MAX_EXPORT_ROWS:
        note_row = last_row + 2
        ws.cell(row=note_row, column=1, value=f"NOTE: Dataset truncated to {MAX_EXPORT_ROWS:,} rows for export safety.")
        ws.cell(row=note_row, column=1).font = Font(italic=True, color="CC0000")
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)


def _build_kpi_sheet(workbook: Workbook, kpis: list[KPIResult]) -> None:
    """Build Sheet 2: KPI Summary."""
    ws = workbook.create_sheet(title="KPI Summary")

    ws["A1"] = "Key Performance Indicators"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")

    if not kpis:
        ws["A3"] = "No KPIs were generated for this dataset."
        ws["A3"].font = Font(italic=True, color="999999")
        return

    # Build DataFrame
    kpi_data = []
    for kpi in kpis:
        kpi_data.append({
            "KPI Name": kpi.name,
            "Value": kpi.formatted_value,
            "Column Used": kpi.column_used or "—",
            "Confidence": kpi.confidence,
            "Description": kpi.description,
        })

    df = pd.DataFrame(kpi_data)
    _write_dataframe_to_sheet(ws, df, start_row=3)
    ws.freeze_panes = "A4"
    _auto_column_widths(ws)


def _build_insights_sheet(workbook: Workbook, insights: list[Insight]) -> None:
    """Build Sheet 3: Business Insights."""
    ws = workbook.create_sheet(title="Business Insights")

    ws["A1"] = "Business Insights"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")

    if not insights:
        ws["A3"] = "No insights were generated for this dataset."
        ws["A3"].font = Font(italic=True, color="999999")
        return

    insight_data = []
    for insight in insights:
        insight_data.append({
            "Observation": insight.observation,
            "Evidence": insight.evidence,
            "Interpretation": insight.business_interpretation,
            "Severity": insight.severity,
            "Category": insight.category,
        })

    df = pd.DataFrame(insight_data)
    _write_dataframe_to_sheet(ws, df, start_row=3)
    ws.freeze_panes = "A4"
    _auto_column_widths(ws)

    # Color-code severity
    severity_col = 4  # Column D
    for row in ws.iter_rows(min_row=4, min_col=severity_col, max_col=severity_col):
        for cell in row:
            if cell.value == "critical":
                cell.fill = CRITICAL_FILL
            elif cell.value == "warning":
                cell.fill = WARNING_FILL
            elif cell.value == "info":
                cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")


def _build_recommendations_sheet(workbook: Workbook, recommendations: list[Recommendation]) -> None:
    """Build Sheet 4: Recommendations."""
    ws = workbook.create_sheet(title="Recommendations")

    ws["A1"] = "Recommendations"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")

    if not recommendations:
        ws["A3"] = "No recommendations were generated for this dataset."
        ws["A3"].font = Font(italic=True, color="999999")
        return

    rec_data = []
    for rec in recommendations:
        rec_data.append({
            "Title": rec.title,
            "Description": rec.description,
            "Evidence": rec.evidence,
            "Priority": rec.priority,
            "Action Category": rec.action_category,
        })

    df = pd.DataFrame(rec_data)
    _write_dataframe_to_sheet(ws, df, start_row=3)
    ws.freeze_panes = "A4"
    _auto_column_widths(ws)

    # Color-code priority
    priority_col = 4  # Column D
    for row in ws.iter_rows(min_row=4, min_col=priority_col, max_col=priority_col):
        for cell in row:
            if cell.value == "high":
                cell.fill = CRITICAL_FILL
            elif cell.value == "medium":
                cell.fill = WARNING_FILL
            elif cell.value == "low":
                cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")


def _build_analyst_report_sheet(workbook: Workbook, report: AnalystReport) -> None:
    """Build Sheet 5: Analyst Report."""
    ws = workbook.create_sheet(title="Analyst Report")

    ws["A1"] = "Analyst Report"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    current_row = 3

    # Dataset Overview
    ws.cell(row=current_row, column=1, value="Dataset Overview")
    ws.cell(row=current_row, column=1).font = SUBTITLE_FONT
    current_row += 1

    overview = report.dataset_overview
    overview_items = [
        ("Total Rows", f"{overview.get('row_count', 0):,}"),
        ("Total Columns", overview.get("column_count", "—")),
        ("Memory Usage", f"{overview.get('memory_usage_mb', 0)} MB"),
        ("Column Names", ", ".join(overview.get("column_names", [])[:10])),
    ]
    for label, value in overview_items:
        ws.cell(row=current_row, column=1, value=label)
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=2, value=value)
        current_row += 1

    current_row += 1

    # Data Quality Summary
    ws.cell(row=current_row, column=1, value="Data Quality Summary")
    ws.cell(row=current_row, column=1).font = SUBTITLE_FONT
    current_row += 1

    quality = report.data_quality_summary
    quality_items = [
        ("Completeness Score", f"{quality.get('completeness_score', 0)}/100"),
        ("Missing Values", f"{quality.get('missing_values', 0):,}"),
        ("Missing Percentage", f"{quality.get('missing_percentage', 0)}%"),
        ("Duplicate Rows", f"{quality.get('duplicate_rows', 0):,}"),
        ("Duplicate Percentage", f"{quality.get('duplicate_percentage', 0)}%"),
        ("Empty Columns", quality.get("empty_columns", 0)),
    ]
    for label, value in quality_items:
        ws.cell(row=current_row, column=1, value=label)
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=2, value=value)
        current_row += 1

    current_row += 1

    # KPI Summary
    ws.cell(row=current_row, column=1, value="KPI Summary")
    ws.cell(row=current_row, column=1).font = SUBTITLE_FONT
    current_row += 1

    if report.kpi_summary:
        kpi_df = pd.DataFrame([
            {"KPI": k.name, "Value": k.formatted_value, "Confidence": k.confidence}
            for k in report.kpi_summary
        ])
        current_row = _write_dataframe_to_sheet(ws, kpi_df, start_row=current_row)
    else:
        ws.cell(row=current_row, column=1, value="No KPIs generated.")
        ws.cell(row=current_row, column=1).font = Font(italic=True, color="999999")
        current_row += 1

    current_row += 1

    # Key Insights
    ws.cell(row=current_row, column=1, value="Key Insights")
    ws.cell(row=current_row, column=1).font = SUBTITLE_FONT
    current_row += 1

    if report.key_insights:
        for insight in report.key_insights:
            ws.cell(row=current_row, column=1, value=insight.observation)
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
            current_row += 1

            ws.cell(row=current_row, column=1, value="Evidence:")
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=9)
            ws.cell(row=current_row, column=2, value=insight.evidence)
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
            current_row += 1

            ws.cell(row=current_row, column=1, value="Interpretation:")
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=9)
            ws.cell(row=current_row, column=2, value=insight.business_interpretation)
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
            current_row += 1

            ws.cell(row=current_row, column=1, value=f"Severity: {insight.severity} | Category: {insight.category}")
            ws.cell(row=current_row, column=1).font = Font(italic=True, size=9, color="666666")
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
            current_row += 1
            current_row += 1
    else:
        ws.cell(row=current_row, column=1, value="No insights generated.")
        ws.cell(row=current_row, column=1).font = Font(italic=True, color="999999")
        current_row += 1

    current_row += 1

    # Recommendations
    ws.cell(row=current_row, column=1, value="Recommendations")
    ws.cell(row=current_row, column=1).font = SUBTITLE_FONT
    current_row += 1

    if report.recommendations:
        for rec in report.recommendations:
            ws.cell(row=current_row, column=1, value=rec.title)
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
            current_row += 1

            ws.cell(row=current_row, column=1, value=rec.description)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
            current_row += 1

            ws.cell(row=current_row, column=1, value=f"Priority: {rec.priority} | Category: {rec.action_category}")
            ws.cell(row=current_row, column=1).font = Font(italic=True, size=9, color="666666")
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
            current_row += 1
            current_row += 1
    else:
        ws.cell(row=current_row, column=1, value="No recommendations generated.")
        ws.cell(row=current_row, column=1).font = Font(italic=True, color="999999")
        current_row += 1

    current_row += 1

    # Cleaning Actions
    ws.cell(row=current_row, column=1, value="Cleaning Actions Performed")
    ws.cell(row=current_row, column=1).font = SUBTITLE_FONT
    current_row += 1

    if report.cleaning_actions:
        clean_df = pd.DataFrame([
            {
                "Timestamp": a.get("timestamp", "—")[:19],
                "Action": a.get("action", "—"),
                "Details": a.get("details", "—"),
                "Affected": a.get("affected_rows", 0),
            }
            for a in report.cleaning_actions
        ])
        current_row = _write_dataframe_to_sheet(ws, clean_df, start_row=current_row)
    else:
        ws.cell(row=current_row, column=1, value="No cleaning actions recorded.")
        ws.cell(row=current_row, column=1).font = Font(italic=True, color="999999")
        current_row += 1

    current_row += 1

    # Analyst Notes
    ws.cell(row=current_row, column=1, value="Analyst Notes")
    ws.cell(row=current_row, column=1).font = SUBTITLE_FONT
    current_row += 1

    ws.cell(row=current_row, column=1, value=report.analyst_notes)
    ws.cell(row=current_row, column=1).fill = NOTE_FILL
    ws.cell(row=current_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 2, end_column=4)
    ws.row_dimensions[current_row].height = 60

    _auto_column_widths(ws)


# =============================================================================
# Public API
# =============================================================================


def generate_analysis_workbook(
    df: pd.DataFrame,
    kpis: list[KPIResult] | None = None,
    insights: list[Insight] | None = None,
    recommendations: list[Recommendation] | None = None,
    report: AnalystReport | None = None,
) -> BytesIO:
    """
    Generate a complete analysis workbook as an in-memory BytesIO buffer.

    Args:
        df: The cleaned dataset.
        kpis: Optional list of KPIResult objects.
        insights: Optional list of Insight objects.
        recommendations: Optional list of Recommendation objects.
        report: Optional AnalystReport object.

    Returns:
        BytesIO buffer containing the .xlsx file.
    """
    workbook = Workbook()

    # Sheet 1: Cleaned Data (always present)
    _build_cleaned_data_sheet(workbook, df)

    # Sheet 2: KPI Summary
    _build_kpi_sheet(workbook, kpis or [])

    # Sheet 3: Business Insights
    _build_insights_sheet(workbook, insights or [])

    # Sheet 4: Recommendations
    _build_recommendations_sheet(workbook, recommendations or [])

    # Sheet 5: Analyst Report
    if report:
        _build_analyst_report_sheet(workbook, report)
    else:
        ws = workbook.create_sheet(title="Analyst Report")
        ws["A1"] = "Analyst Report"
        ws["A1"].font = TITLE_FONT
        ws["A3"] = "No analyst report was generated. Generate a report in the Business Intelligence Center first."
        ws["A3"].font = Font(italic=True, color="999999")
        ws.merge_cells("A3:D3")
        _auto_column_widths(ws)

    # Save to buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer