"""
Multi-File Comparison Engine — Version 1
========================================

Compare two datasets to identify added, removed, and modified records.
Expandable architecture for future comparison enhancements.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from io import BytesIO
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


class ComparisonServiceError(Exception):
    """Raised when a comparison operation fails."""


@dataclass
class ComparisonResult:
    """Result of a dataset comparison."""
    summary: dict[str, Any]
    added_records: pd.DataFrame
    removed_records: pd.DataFrame
    modified_records: pd.DataFrame
    numeric_comparison: pd.DataFrame
    join_key: str
    dataset_a_rows: int
    dataset_b_rows: int
    matched_records: int
    added_count: int
    removed_count: int
    modified_count: int


# =============================================================================
# Internal Helpers
# =============================================================================

def _validate_join_key(df: pd.DataFrame, join_key: str, label: str) -> None:
    """Validate that a join key exists and is unique."""
    if not join_key or not join_key.strip():
        raise ComparisonServiceError("Join key cannot be empty.")
    if join_key not in df.columns:
        raise ComparisonServiceError(f"Join key '{join_key}' not found in {label}.")
    if df[join_key].isna().all():
        raise ComparisonServiceError(f"Join key '{join_key}' contains only null values in {label}.")
    dupes = df[join_key].duplicated().sum()
    if dupes > 0:
        raise ComparisonServiceError(
            f"Join key '{join_key}' has {dupes:,} duplicate(s) in {label}. "
            "Please use a unique key or de-duplicate first."
        )


def _get_common_columns(df_a: pd.DataFrame, df_b: pd.DataFrame, join_key: str) -> list[str]:
    """Return common columns between two dataframes, excluding the join key."""
    cols_a = set(df_a.columns)
    cols_b = set(df_b.columns)
    common = cols_a & cols_b
    if join_key not in common:
        raise ComparisonServiceError(f"Join key '{join_key}' must exist in both datasets.")
    return sorted([c for c in common if c != join_key])


# =============================================================================
# Public API
# =============================================================================

def find_added_records(df_a: pd.DataFrame, df_b: pd.DataFrame, join_key: str) -> pd.DataFrame:
    """Records present in Dataset B but not in Dataset A."""
    keys_a = set(df_a[join_key].astype(str))
    keys_b = set(df_b[join_key].astype(str))
    added_keys = keys_b - keys_a
    return df_b[df_b[join_key].astype(str).isin(added_keys)].copy()


def find_removed_records(df_a: pd.DataFrame, df_b: pd.DataFrame, join_key: str) -> pd.DataFrame:
    """Records present in Dataset A but not in Dataset B."""
    keys_a = set(df_a[join_key].astype(str))
    keys_b = set(df_b[join_key].astype(str))
    removed_keys = keys_a - keys_b
    return df_a[df_a[join_key].astype(str).isin(removed_keys)].copy()


def find_modified_records(
    df_a: pd.DataFrame,
    df_b: pd.DataFrame,
    join_key: str,
) -> pd.DataFrame:
    """
    Records with the same join key but different values in common columns.
    Returns a long-format DataFrame: [Join Key, Column, Old Value, New Value].
    """
    common_cols = _get_common_columns(df_a, df_b, join_key)
    keys_a = set(df_a[join_key].astype(str))
    keys_b = set(df_b[join_key].astype(str))
    common_keys = keys_a & keys_b

    df_a_common = df_a[df_a[join_key].astype(str).isin(common_keys)].copy()
    df_b_common = df_b[df_b[join_key].astype(str).isin(common_keys)].copy()

    # Normalize join key type for reliable merging
    df_a_common[join_key] = df_a_common[join_key].astype(str)
    df_b_common[join_key] = df_b_common[join_key].astype(str)

    merged = df_a_common.merge(
        df_b_common,
        on=join_key,
        how="inner",
        suffixes=("_A", "_B"),
    )

    modified_rows: list[pd.DataFrame] = []
    for col in common_cols:
        col_a = f"{col}_A"
        col_b = f"{col}_B"
        if col_a in merged.columns and col_b in merged.columns:
            # NaN-safe comparison
            mask = merged[col_a].fillna("___NULL___") != merged[col_b].fillna("___NULL___")
            if mask.any():
                changed = merged[mask][[join_key, col_a, col_b]].copy()
                changed = changed.rename(columns={col_a: "Old Value", col_b: "New Value"})
                changed["Column"] = col
                changed = changed[[join_key, "Column", "Old Value", "New Value"]]
                modified_rows.append(changed)

    if not modified_rows:
        return pd.DataFrame(columns=[join_key, "Column", "Old Value", "New Value"])

    return pd.concat(modified_rows, ignore_index=True)


def generate_numeric_comparison(
    df_a: pd.DataFrame,
    df_b: pd.DataFrame,
    common_keys: set,
    join_key: str,
) -> pd.DataFrame:
    """Compare numeric column totals between matching records."""
    if not common_keys:
        return pd.DataFrame()

    df_a_common = df_a[df_a[join_key].astype(str).isin(common_keys)].copy()
    df_b_common = df_b[df_b[join_key].astype(str).isin(common_keys)].copy()

    numeric_cols_a = df_a_common.select_dtypes(include="number").columns.tolist()
    numeric_cols_b = df_b_common.select_dtypes(include="number").columns.tolist()
    numeric_cols = sorted(list(set(numeric_cols_a) & set(numeric_cols_b)))
    numeric_cols = [c for c in numeric_cols if c != join_key]

    if not numeric_cols:
        return pd.DataFrame()

    rows: list[dict[str, Any]] = []
    for col in numeric_cols:
        total_a = pd.to_numeric(df_a_common[col], errors="coerce").sum()
        total_b = pd.to_numeric(df_b_common[col], errors="coerce").sum()
        diff = total_b - total_a
        if total_a != 0:
            pct_change = (diff / total_a) * 100
            pct_str = f"{pct_change:.2f}%"
        elif diff == 0:
            pct_str = "0.00%"
        else:
            pct_str = "N/A"
        rows.append({
            "Column": col,
            "Total A": total_a,
            "Total B": total_b,
            "Difference": diff,
            "Percent Change": pct_str,
        })

    return pd.DataFrame(rows)


def generate_comparison_summary(result: ComparisonResult) -> dict[str, Any]:
    """Return a formatted summary dictionary from a ComparisonResult."""
    return result.summary


def analyze_datasets(
    df_a: pd.DataFrame,
    df_b: pd.DataFrame,
    join_key: str,
) -> ComparisonResult:
    """
    Full comparison analysis between two datasets.
    """
    if df_a is None or df_a.empty:
        raise ComparisonServiceError("Primary dataset (A) is empty.")
    if df_b is None or df_b.empty:
        raise ComparisonServiceError("Comparison dataset (B) is empty.")

    _validate_join_key(df_a, join_key, "primary dataset")
    _validate_join_key(df_b, join_key, "comparison dataset")

    added_records = find_added_records(df_a, df_b, join_key)
    removed_records = find_removed_records(df_a, df_b, join_key)
    modified_records = find_modified_records(df_a, df_b, join_key)

    keys_a = set(df_a[join_key].astype(str))
    keys_b = set(df_b[join_key].astype(str))
    common_keys = keys_a & keys_b

    numeric_comparison = generate_numeric_comparison(df_a, df_b, common_keys, join_key)
    common_cols = _get_common_columns(df_a, df_b, join_key)

    summary = {
        "dataset_a_rows": len(df_a),
        "dataset_b_rows": len(df_b),
        "matched_records": len(common_keys),
        "added_records": len(added_records),
        "removed_records": len(removed_records),
        "modified_records": len(modified_records),
        "common_columns": len(common_cols),
        "numeric_columns_compared": len(numeric_comparison) if not numeric_comparison.empty else 0,
    }

    return ComparisonResult(
        summary=summary,
        added_records=added_records,
        removed_records=removed_records,
        modified_records=modified_records,
        numeric_comparison=numeric_comparison,
        join_key=join_key,
        dataset_a_rows=len(df_a),
        dataset_b_rows=len(df_b),
        matched_records=len(common_keys),
        added_count=len(added_records),
        removed_count=len(removed_records),
        modified_count=len(modified_records),
    )


# =============================================================================
# Export
# =============================================================================

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_DATA_ALIGNMENT = Alignment(vertical="center", wrap_text=False)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_TITLE_FONT = Font(bold=True, size=14, color="366092")
_SUBTITLE_FONT = Font(bold=True, size=11, color="000000")


def _auto_column_widths(worksheet) -> None:
    """Auto-adjust column widths based on content."""
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 60)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def _apply_header_style(worksheet, row_idx: int = 1) -> None:
    """Apply header styling to a specific row."""
    for cell in worksheet[row_idx]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER


def _apply_data_borders(worksheet, start_row: int = 2) -> None:
    """Apply thin borders to all data cells."""
    for row in worksheet.iter_rows(min_row=start_row):
        for cell in row:
            cell.border = _THIN_BORDER
            cell.alignment = _DATA_ALIGNMENT


def _write_df_to_sheet(
    worksheet,
    df: pd.DataFrame,
    start_row: int = 1,
    start_col: int = 1,
) -> int:
    """Write a DataFrame to a worksheet and return the last row index."""
    if df.empty:
        ws_cell = worksheet.cell(row=start_row, column=start_col, value="No data available.")
        ws_cell.font = Font(italic=True, color="999999")
        return start_row

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, start_col):
            worksheet.cell(row=r_idx, column=c_idx, value=value)

    _apply_header_style(worksheet, row_idx=start_row)
    _apply_data_borders(worksheet, start_row=start_row + 1)

    return start_row + len(df)


def generate_comparison_workbook(result: ComparisonResult) -> BytesIO:
    """
    Generate a comparison report workbook as an in-memory BytesIO buffer.
    Sheets: Summary, Added Records, Removed Records, Modified Records, Numeric Differences.
    """
    wb = Workbook()

    # Sheet 1: Summary
    ws = wb.active
    ws.title = "Summary"

    ws["A1"] = "Multi-File Comparison Report"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells("A1:D1")

    ws["A2"] = f"Join Key: {result.join_key}"
    ws["A2"].font = _SUBTITLE_FONT
    ws.merge_cells("A2:D2")

    summary_data = [
        ["Metric", "Value"],
        ["Rows in Dataset A", f"{result.dataset_a_rows:,}"],
        ["Rows in Dataset B", f"{result.dataset_b_rows:,}"],
        ["Matched Records", f"{result.matched_records:,}"],
        ["Added Records", f"{result.added_count:,}"],
        ["Removed Records", f"{result.removed_count:,}"],
        ["Modified Records", f"{result.modified_count:,}"],
        ["Common Columns (excl. key)", result.summary.get("common_columns", 0)],
        ["Numeric Columns Compared", result.summary.get("numeric_columns_compared", 0)],
    ]

    for r_idx, row in enumerate(summary_data, 3):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 3:
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL
                cell.alignment = _HEADER_ALIGNMENT
            cell.border = _THIN_BORDER

    _auto_column_widths(ws)

    # Sheet 2: Added Records
    ws2 = wb.create_sheet(title="Added Records")
    _write_df_to_sheet(ws2, result.added_records)

    # Sheet 3: Removed Records
    ws3 = wb.create_sheet(title="Removed Records")
    _write_df_to_sheet(ws3, result.removed_records)

    # Sheet 4: Modified Records
    ws4 = wb.create_sheet(title="Modified Records")
    _write_df_to_sheet(ws4, result.modified_records)

    # Sheet 5: Numeric Differences
    ws5 = wb.create_sheet(title="Numeric Differences")
    _write_df_to_sheet(ws5, result.numeric_comparison)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer